import os
import numpy as np
import pandas as pd
from sqlalchemy import create_engine, text
from openpyxl.formatting.rule import ColorScaleRule
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

DB_URL = os.getenv("DB_URL", "postgresql+psycopg2://postgres:postgres@db:5432/postgres")
EXPORTS_DIR = os.getenv("EXPORTS_DIR", "exports")
os.makedirs(EXPORTS_DIR, exist_ok=True)

engine = create_engine(DB_URL)

# ---- DataFrames from SQL (2+ JOINs) ----
sql_city = '''
SELECT c.customer_city AS city, COUNT(DISTINCT o.order_id) AS orders_count
FROM orders o
JOIN customers c ON c.customer_id = o.customer_id
JOIN order_items oi ON oi.order_id = o.order_id
WHERE o.order_status IN ('delivered','shipped','invoiced')
GROUP BY city
ORDER BY orders_count DESC
LIMIT 30
'''
df_city = pd.read_sql_query(text(sql_city), engine)

sql_month = '''
SELECT DATE_TRUNC('month', o.order_purchase_timestamp) AS month,
       SUM(oi.price + oi.freight_value) AS revenue
FROM orders o
JOIN order_items oi ON oi.order_id = o.order_id
WHERE o.order_status IN ('delivered','shipped','invoiced')
GROUP BY month
ORDER BY month
'''
df_month = pd.read_sql_query(text(sql_month), engine)
df_month['month'] = pd.to_datetime(df_month['month'])

sql_cat = '''
SELECT COALESCE(p.product_category_name_english,'(unknown)') AS category,
       COUNT(*) AS items_cnt,
       ROUND(AVG(oi.price)::numeric, 2) AS avg_price,
       ROUND(SUM(oi.price)::numeric, 2) AS revenue
FROM orders o
JOIN order_items oi ON oi.order_id = o.order_id
JOIN products p ON p.product_id = oi.product_id
WHERE o.order_status IN ('delivered','shipped','invoiced')
GROUP BY category
HAVING COUNT(*) >= 50
ORDER BY revenue DESC
LIMIT 25
'''
df_cat = pd.read_sql_query(text(sql_cat), engine)

export_data = {
    "Top Cities by Orders": df_city,
    "Monthly Revenue": df_month,
    "Category Summary": df_cat
}

out_xlsx = os.path.join(EXPORTS_DIR, "analytics_report.xlsx")

with pd.ExcelWriter(out_xlsx, engine="openpyxl") as writer:
    for sheet, df in export_data.items():
        df.to_excel(writer, sheet_name=sheet, index=False, startrow=1, startcol=1)
        ws = writer.sheets[sheet]

        # Title & header styling
        ws["B1"] = f"Report: {sheet}"
        ws["B1"].font = Font(size=16, bold=True, color="1F4E79")
        ws.freeze_panes = "B3"

        header_fill = PatternFill(start_color="D9E2F3", end_color="D9E2F3", fill_type="solid")
        header_font = Font(bold=True, color="1F4E79")

        for col_num in range(2, len(df.columns) + 2):
            cell = ws.cell(row=2, column=col_num)
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal="center", vertical="center")

        # Auto width
        for column in ws.columns:
            max_len = 0
            col_letter = column[0].column_letter
            for cell in column:
                try:
                    max_len = max(max_len, len(str(cell.value)))
                except:
                    pass
            ws.column_dimensions[col_letter].width = min(max_len + 2, 50)

        # Conditional formatting for numeric columns
        numeric_cols = df.select_dtypes(include=[np.number]).columns.tolist()
        for idx, col_name in enumerate(df.columns, start=2):
            if col_name in numeric_cols and len(df) > 1:
                col_letter = ws.cell(row=1, column=idx).column_letter
                addr = f"{col_letter}3:{col_letter}{len(df)+2}"
                rule = ColorScaleRule(
                    start_type="min", start_color="63BE7B",
                    mid_type="percentile", mid_value=50, mid_color="FFDD71",
                    end_type="max", end_color="F8696B"
                )
                ws.conditional_formatting.add(addr, rule)

        # AutoFilter
        ws.auto_filter.ref = f"B2:{ws.cell(row=len(df)+2, column=len(df.columns)+1).coordinate}"

        # Thin border around table
        thin = Border(left=Side(style="thin"), right=Side(style="thin"),
                      top=Side(style="thin"), bottom=Side(style="thin"))
        for r in range(2, len(df)+3):
            for c in range(2, len(df.columns)+2):
                ws.cell(row=r, column=c).border = thin

total_rows = sum(len(df) for df in export_data.values())
print(f"Created file {out_xlsx}, {len(export_data)} sheets, {total_rows} rows")
